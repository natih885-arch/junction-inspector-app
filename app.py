import streamlit as st
import pandas as pd
import io
import base64
from PIL import Image, ImageDraw
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from streamlit_drawable_canvas import st_canvas

# ---------------------------------------------------------
# הגדרות תצורה ועיצוב דף Streamlit
# ---------------------------------------------------------
st.set_page_config(
    page_title='מערכת פיקוח תשתיות רמזורים ובקרה',
    page_icon='🚦',
    layout='wide',
    initial_sidebar_state='expanded'
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Rubik:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] {
    font-family: 'Rubik', sans-serif;
    direction: rtl;
}
.main-header {
    background: linear-gradient(135deg, #0f2b48 0%, #1e4d7b 100%);
    color: white;
    padding: 24px 30px;
    border-radius: 12px;
    box-shadow: 0 8px 20px rgba(15, 43, 72, 0.15);
    margin-bottom: 25px;
    border-right: 8px solid #00a887;
}
.main-header h1 { font-size: 28px; font-weight: 700; margin: 0; color: #ffffff; }
.main-header p { font-size: 15px; margin-top: 6px; margin-bottom: 0; color: #d0e1f9; }
.stButton>button {
    width: 100%;
    background: linear-gradient(90deg, #0f2b48 0%, #1e4d7b 100%);
    color: white;
    font-weight: 700;
    border: none;
    padding: 12px 24px;
    border-radius: 8px;
    transition: all 0.3s ease;
}
.stButton>button:hover {
    background: linear-gradient(90deg, #1e4d7b 0%, #00a887 100%);
    box-shadow: 0 4px 12px rgba(0, 168, 135, 0.3);
}
.legend-chip {
    display: inline-block;
    padding: 4px 10px;
    border-radius: 14px;
    color: white;
    font-size: 12px;
    font-weight: 600;
    margin-inline-end: 6px;
    margin-bottom: 6px;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
<h1>🚦 מערכת פיקוח הנדסית, בקרה וכתב כמויות - תשתיות צמתים</h1>
<p>מערכת פיקוח מקצועית להנפקת דוחות מנכ"ל | עריכת תוכניות עבודה, בדיקות הארקה וחישובי כתבי כמויות כבלים</p>
</div>
""", unsafe_allow_html=True)

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "🎨 סקיצת תוואי ותוכנית עבודה",
    "🔢 מפרט ציוד וספירת שטח",
    "⚡ פרוטוקול הארקה ובטיחות",
    "📊 כתב כמויות וחישוב כבלים",
    "📥 הפקת דוח מנהלים (Excel מקצועי)"
])

# ---------------------------------------------------------
# בנק האלמנטים
# ---------------------------------------------------------
ELEMENT_BANK = {
    "pole_metal":     {"label": "עמוד מתכת",        "emoji": "🗼", "color": "#7f8c8d"},
    "pole_concrete":  {"label": "עמוד בטון",         "emoji": "🧱", "color": "#8d6e63"},
    "car_light":      {"label": "פנס רכב",           "emoji": "🚦", "color": "#e74c3c"},
    "ped_light":      {"label": "פנס הולכי רגל",     "emoji": "🚶", "color": "#3498db"},
    "blinker":        {"label": "בלינקר / מהבהב",    "emoji": "🔶", "color": "#f39c12"},
    "bike_light":     {"label": "פנס אופניים",       "emoji": "🚲", "color": "#9b59b6"},
    "sign":           {"label": "תמרור",             "emoji": "🛑", "color": "#f1c40f"},
    "cabinet":        {"label": "ארון פיקוד",        "emoji": "🗄️", "color": "#2c3e50"},
}

BANK_TO_AFTER_KEY = {
    "pole_metal": "poles_metal_after",
    "pole_concrete": "poles_concrete_after",
    "car_light": "car_lights_after",
    "ped_light": "ped_lights_after",
    "blinker": "blinkers_after",
    "bike_light": "bike_lights_after",
    "sign": "signs_after",
    "cabinet": "cabinets_after",
}

DEFAULTS = {
    "poles_metal_before": 4, "poles_metal_after": 6,
    "poles_concrete_before": 2, "poles_concrete_after": 0,
    "cabinets_before": 1, "cabinets_after": 1,
    "car_lights_before": 6, "car_lights_after": 8,
    "ped_lights_before": 4, "ped_lights_after": 6,
    "blinkers_before": 1, "blinkers_after": 3,
    "bike_lights_before": 0, "bike_lights_after": 2,
    "signs_before": 3, "signs_after": 3,
}

for _k, _v in DEFAULTS.items():
    st.session_state.setdefault(_k, _v)

st.session_state.setdefault("canvas_reset_counter", 0)

def create_junction_background(junction_type="4_way"):
    img = Image.new("RGB", (900, 500), "#2c3e50")
    draw = ImageDraw.Draw(img)
    if junction_type == "4_way":
        draw.rectangle([350, 0, 550, 500], fill="#34495e")
        draw.rectangle([0, 175, 900, 325], fill="#34495e")
        for y in range(180, 320, 20):
            draw.line([(330, y), (350, y)], fill="white", width=4)
            draw.line([(550, y), (570, y)], fill="white", width=4)
        for x in range(355, 545, 20):
            draw.line([(x, 155), (x, 175)], fill="white", width=4)
            draw.line([(x, 325), (x, 345)], fill="white", width=4)
    else:  # 3_way / T-Junction
        draw.rectangle([0, 175, 900, 325], fill="#34495e")
        draw.rectangle([350, 175, 550, 500], fill="#34495e")
        for y in range(180, 320, 20):
            draw.line([(330, y), (350, y)], fill="white", width=4)
            draw.line([(550, y), (570, y)], fill="white", width=4)
        for x in range(355, 545, 20):
            draw.line([(x, 325), (x, 345)], fill="white", width=4)
    return img

def pil_to_base64(img):
    """ממיר תמונת PIL למחרוזת Base64 כדי למנוע את השגיאה ב-streamlit_drawable_canvas"""
    buffered = io.BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode()
    return f"data:image/png;base64,{img_str}"

def count_canvas_elements(json_data):
    counts = {k: 0 for k in ELEMENT_BANK}
    if not json_data:
        return counts
    color_to_key = {v["color"].lower(): k for k, v in ELEMENT_BANK.items()}
    for obj in json_data.get("objects", []):
        fill = str(obj.get("fill") or "").lower()
        if fill in color_to_key:
            counts[color_to_key[fill]] += 1
    return counts

# ---------------------------------------------------------
# כרטיסייה 1: שרטוט ובנק פנסים
# ---------------------------------------------------------
with tab1:
    st.subheader("📌 פרטי מפתח לסיור הפיקוח")
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        junction_name = st.text_input("שם / מספר הצומת והפרויקט", "אלנבי - מנחם בגין (תוואי קו סגול)")
        inspector_name = st.text_input("שם המפקח האחראי", "נתנאל הררי - מפקח תשתיות מוסמך")
        contractor_name = st.text_input("קבלן מבצע / חברה מבצעת", "חברה מבצעת - אגף תשתיות")

    with col_b:
        inspection_date = st.date_input("תאריך סיור הפיקוח")
        project_phase = st.selectbox("שלב הסדר התנועה (תמ''ת/תמ''ק)", [
            "שלב א' - מצב קיים (לפני שינוי)",
            "שלב ב' - הסדר זמני (כבילה עילית/מעקף)",
            "שלב ג' - הסדר קבוע (כבילה תחתית / סופי)"
        ])
        cabling_mode = st.radio("סוג כבילה מרכזי בתוואי",
                                 ["כבילה תחתית (שרוולים/צנרת באדמה)", "כבילה עילית (זמנית על מתוחים)"])

    with col_c:
        general_notes = st.text_area(
            "הערות הנדסיות ודגשי שטח",
            "תוואי כבילה מתוכנן מצד מזרח למערב כולל מעבר ארון פיקוד. בוצעה בדיקת שרוולים והכנה לביטון.",
            height=128
        )

    st.divider()
    st.subheader("🗺️ בחירת בסיס תרשים הצומת")
    bg_option = st.radio(
        "בחר מקור לרקע הצומת:",
        ["מחולל צומת תלת-ממדי/דו-ממדי מובנה", "העלאת תוכנית / אורתופוטו (קובץ תמונה)"],
        horizontal=True
    )

    bg_img = None
    if bg_option == "מחולל צומת תלת-ממדי/דו-ממדי מובנה":
        j_type = st.selectbox("בחר סוג צומת:", ["צומת 4 כיוונים (X)", "צומת 3 כיוונים (T)"])
        j_code = "4_way" if "4" in j_type else "3_way"
        bg_img = create_junction_background(j_code)
        st.session_state["active_bg_img"] = bg_img
    else:
        uploaded_bg = st.file_uploader("בחרו קובץ תמונה של הצומת (PNG/JPG):", type=["png", "jpg", "jpeg"])
        if uploaded_bg is not None:
            bg_img = Image.open(uploaded_bg).convert("RGB").resize((900, 500))
            st.session_state["active_bg_img"] = bg_img
        elif "active_bg_img" in st.session_state:
            bg_img = st.session_state["active_bg_img"]

    st.divider()
    st.subheader("🚥 בנק פנסים, תמרורים וסרגל כלי שרטוט")
    st.caption(
        "בחרו 'הוספת אלמנט מבנק' כדי להטביע פנס/עמוד/ציוד על השרטוט בלחיצה, "
        "ואז עברו למצב '🖐️ הזזה ועריכה' כדי לגרור ולמקם כל אלמנט בדיוק במקום שבו הוא ניצב בשטח."
    )

    mode_choice = st.radio(
        "מצב עבודה בסקיצה:",
        [
            "🖐️ הזזה ועריכת אלמנטים",
            "➕ הוספת אלמנט מבנק הציוד",
            "✏️ ציור תוואי כבלים (חופשי)",
            "📏 קו ישר (שרוול/חפירה)",
            "🟦 סימון שטח / ארון פיקוד (מלבן)",
        ],
        horizontal=True,
    )

    col_t1, col_t2, col_t3 = st.columns(3)
    selected_elem_key = None
    if mode_choice == "➕ הוספת אלמנט מבנק הציוד":
        with col_t1:
            selected_elem_key = st.selectbox(
                "בחר אלמנט להצבה:",
                list(ELEMENT_BANK.keys()),
                format_func=lambda k: f"{ELEMENT_BANK[k]['emoji']} {ELEMENT_BANK[k]['label']}",
            )
        elem = ELEMENT_BANK[selected_elem_key]
        drawing_mode = "point"
        stroke_color = elem["color"]
        fill_color = elem["color"]
        with col_t2:
            st.color_picker("צבע האלמנט (קבוע לפי הבנק):", elem["color"], disabled=True)
        with col_t3:
            stroke_width = st.slider("גודל האלמנט על השרטוט:", 6, 30, 14)
        point_radius = stroke_width
    else:
        mode_map = {
            "🖐️ הזזה ועריכת אלמנטים": "transform",
            "✏️ ציור תוואי כבלים (חופשי)": "freedraw",
            "📏 קו ישר (שרוול/חפירה)": "line",
            "🟦 סימון שטח / ארון פיקוד (מלבן)": "rect",
        }
        drawing_mode = mode_map[mode_choice]
        with col_t1:
            st.info(f"מצב פעיל: {mode_choice}")
        with col_t2:
            stroke_color = st.color_picker("צבע התוואי/האלמנט:", "#00A887")
        with col_t3:
            stroke_width = st.slider("עובי הקו:", 1, 15, 4)
        fill_color = "rgba(0, 168, 135, 0.4)"
        point_radius = 3

    col_reset, _ = st.columns([1, 3])
    with col_reset:
        if st.button("🗑️ נקה שרטוט (איפוס הסקיצה)"):
            st.session_state["canvas_reset_counter"] += 1
            st.session_state.pop("canvas_sketch_image", None)

    legend_html = "".join(
        f'<span class="legend-chip" style="background:{v["color"]}">{v["emoji"]} {v["label"]}</span>'
        for v in ELEMENT_BANK.values()
    )
    st.markdown(legend_html, unsafe_allow_html=True)

    # המרה ל-Base64 כדי למנוע את שגיאת ה-AttributeError
    bg_image_url = pil_to_base64(bg_img) if bg_img is not None else ""

    canvas_key = f"canvas_junction_interactive_{st.session_state['canvas_reset_counter']}"
    canvas_result = st_canvas(
        fill_color=fill_color,
        stroke_width=stroke_width,
        stroke_color=stroke_color,
        background_image=None,
        background_color="#FFFFFF" if bg_img is None else "",
        background_image_url=bg_image_url if bg_img is not None else None,
        height=500,
        width=900,
        drawing_mode=drawing_mode,
        point_display_radius=point_radius,
        update_streamlit=True,
        key=canvas_key,
    )

    if canvas_result.image_data is not None:
        st.session_state["canvas_sketch_image"] = canvas_result.image_data

    sketch_counts = count_canvas_elements(canvas_result.json_data)
    st.session_state["sketch_counts"] = sketch_counts

    st.markdown("#### 🧮 ספירה אוטומטית מתוך הסקיצה")
    count_cols = st.columns(len(ELEMENT_BANK))
    for col, (k, v) in zip(count_cols, ELEMENT_BANK.items()):
        col.metric(f"{v['emoji']} {v['label']}", sketch_counts.get(k, 0))

    if st.button("✅ עדכן את כתב הכמויות (טאב 2) לפי מה שמופיע בסקיצה"):
        for bank_key, session_key in BANK_TO_AFTER_KEY.items():
            st.session_state[session_key] = sketch_counts.get(bank_key, 0)
        st.success("כתב הכמויות בטאב 'מפרט ציוד וספירת שטח' עודכן לפי הסקיצה בפועל.")

# ---------------------------------------------------------
# כרטיסייה 2: ספירת ציוד בשטח
# ---------------------------------------------------------
with tab2:
    st.subheader("🔢 תיעוד וספירת ציוד בצומת (לפני / אחרי)")
    st.caption("ניתן למלא ידנית, או להשתמש בכפתור 'עדכן את כתב הכמויות' בטאב הסקיצה כדי למלא אוטומטית את עמודת 'אחרי'.")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("### 🏛️ עמודים, ארונות ותשתיות")
        poles_metal_before = st.number_input("עמודי מתכת (לפני)", min_value=0, key="poles_metal_before")
        poles_metal_after = st.number_input("עמודי מתכת (אחרי)", min_value=0, key="poles_metal_after")
        poles_concrete_before = st.number_input("עמודי בטון (לפני)", min_value=0, key="poles_concrete_before")
        poles_concrete_after = st.number_input("עמודי בטון (אחרי)", min_value=0, key="poles_concrete_after")
        cabinets_before = st.number_input("ארונות פיקוד (לפני)", min_value=0, key="cabinets_before")
        cabinets_after = st.number_input("ארונות פיקוד (אחרי)", min_value=0, key="cabinets_after")

    with col2:
        st.markdown("### 🚥 פנסי תנועה ופיקוד")
        car_lights_before = st.number_input("פנסי רכב (לפני)", min_value=0, key="car_lights_before")
        car_lights_after = st.number_input("פנסי רכב (אחרי)", min_value=0, key="car_lights_after")
        ped_lights_before = st.number_input("פנסי הולכי רגל (לפני)", min_value=0, key="ped_lights_before")
        ped_lights_after = st.number_input("פנסי הולכי רגל (אחרי)", min_value=0, key="ped_lights_after")

    with col3:
        st.markdown("### 🚲 בלינקרים, אופניים ותמרורים")
        blinkers_before = st.number_input("בלינקרים / מהבהבים (לפני)", min_value=0, key="blinkers_before")
        blinkers_after = st.number_input("בלינקרים / מהבהבים (אחרי)", min_value=0, key="blinkers_after")
        bike_lights_before = st.number_input("פנסי אופניים (לפני)", min_value=0, key="bike_lights_before")
        bike_lights_after = st.number_input("פנסי אופניים (אחרי)", min_value=0, key="bike_lights_after")
        signs_before = st.number_input("תמרורים (לפני)", min_value=0, key="signs_before")
        signs_after = st.number_input("תמרורים (אחרי)", min_value=0, key="signs_after")

# ---------------------------------------------------------
# כרטיסייה 3: אישור הארקה ובטיחות
# ---------------------------------------------------------
with tab3:
    st.subheader("⚡ צ'קליסט ובדיקות בטיחות הארקה בחשמל")
    col_g1, col_g2 = st.columns(2)
    with col_g1:
        g_poles = st.checkbox("✅ בוצעה בדיקת חיבור הארקה לכל עמודי הצומת", value=True)
        g_cabinet = st.checkbox("✅ הארקת ארון פיקוד מרכזי מחוברת ותקינה", value=True)
        g_continuity = st.checkbox("✅ בוצעה בדיקת רציפות הארקה (Continuity Test)", value=True)
        g_value = st.number_input("ערך הארקה שנמדד (אוהם Ω)", min_value=0.0, value=1.2, step=0.1)

    with col_g2:
        grounding_img = st.file_uploader("📷 צילום פתח עמוד / פס הארקה", type=["jpg", "png", "jpeg"])
        if grounding_img:
            image = Image.open(grounding_img)
            st.image(image, caption="תמונת הארקה מהשטח", width=300)

    is_grounding_approved = g_poles and g_cabinet and g_continuity
    if is_grounding_approved:
        st.success("🛡️ סטטוס בטיחות: הארקת הצומת מאושרת ותקינה על פי תקנות החשמל!")
    else:
        st.error("❌ אזהרה הנדסית: טרם הושלמו כל בדיקות ההארקה הנדרשות בצומת!")

# ---------------------------------------------------------
# כרטיסייה 4: כתב כמויות וחישוב כבלים
# ---------------------------------------------------------
with tab4:
    st.subheader("📊 כתב כמויות הנדסי וניתוח שינויים (Δ)")
    boq_data = [
        {"תיאור הרכיב": "עמודי מתכת (זרוע/ישר)", "לפני": poles_metal_before, "אחרי": poles_metal_after},
        {"תיאור הרכיב": "עמודי בטון", "לפני": poles_concrete_before, "אחרי": poles_concrete_after},
        {"תיאור הרכיב": "ארונות פיקוד", "לפני": cabinets_before, "אחרי": cabinets_after},
        {"תיאור הרכיב": "פנסי תנועה לרכב", "לפני": car_lights_before, "אחרי": car_lights_after},
        {"תיאור הרכיב": "פנסי הולכי רגל", "לפני": ped_lights_before, "אחרי": ped_lights_after},
        {"תיאור הרכיב": "בלינקרים / מהבהבים", "לפני": blinkers_before, "אחרי": blinkers_after},
        {"תיאור הרכיב": "פנסי אופניים", "לפני": bike_lights_before, "אחרי": bike_lights_after},
        {"תיאור הרכיב": "תמרורים", "לפני": signs_before, "אחרי": signs_after},
    ]

    df_boq = pd.DataFrame(boq_data)
    df_boq["שינוי (דלתא Δ)"] = df_boq["אחרי"] - df_boq["לפני"]
    st.dataframe(df_boq, use_container_width=True)

    added_car_lights = max(0, car_lights_after - car_lights_before)
    added_ped_lights = max(0, ped_lights_after - ped_lights_before)
    added_blinkers = max(0, blinkers_after - blinkers_before)

    cable_heavy = (added_car_lights * 32) + (added_blinkers * 20)
    cable_light = added_ped_lights * 22
    cable_ground = (poles_metal_after * 6) + 15

    st.divider()
    st.markdown("### 🔌 תחשיב אומדן כבילה מתוכנן")
    m1, m2, m3 = st.columns(3)
    m1.metric("כבל פיקוד רמזורים כבד (NYCWY / NYY)", f"{cable_heavy} מטר")
    m2.metric("כבל פיקוד הולכי רגל", f"{cable_light} מטר")
    m3.metric("כבל הארקה תקני (חשוף / בידוד)", f"{cable_ground} מטר")

# ---------------------------------------------------------
# כרטיסייה 5: הפקת דוח Excel הנדסי ומעוצב
# ---------------------------------------------------------
with tab5:
    st.subheader("📥 הפקת דוח מנהלים הנדסי (Excel)")
    st.markdown("הדוח כולל עיצוב מותאם, גיליון תקציר מנהלים, כתב כמויות, פרוטוקול הארקה וסקיצה הנדסית משולבת.")

    def generate_excel_report():
        output = io.BytesIO()
        wb = openpyxl.Workbook()

        COLOR_PRIMARY = "0F2B48"
        COLOR_SECONDARY = "1E4D7B"
        COLOR_CARD_BG = "EBF2F8"
        COLOR_LIGHT_BG = "F4F7FA"
        COLOR_ZEBRA = "F9FBFD"
        COLOR_BORDER = "D1D5DB"
        COLOR_SUCCESS_BG = "E6F4EA"
        COLOR_SUCCESS_FG = "137333"
        COLOR_DANGER_BG = "FCE8E6"
        COLOR_DANGER_FG = "C5221F"

        font_main_title = Font(name="Arial", size=18, bold=True, color="FFFFFF")
        font_sub_title = Font(name="Arial", size=11, bold=False, color="D0E1F9")
        font_section_header = Font(name="Arial", size=13, bold=True, color=COLOR_PRIMARY)
        font_tbl_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Arial", size=10, color="000000")
        font_data_bold = Font(name="Arial", size=10, bold=True, color="000000")
        font_metric_num = Font(name="Arial", size=16, bold=True, color=COLOR_PRIMARY)
        font_metric_lbl = Font(name="Arial", size=9, bold=True, color="555555")

        fill_primary = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
        fill_secondary = PatternFill(start_color=COLOR_SECONDARY, end_color=COLOR_SECONDARY, fill_type="solid")
        fill_zebra = PatternFill(start_color=COLOR_ZEBRA, end_color=COLOR_ZEBRA, fill_type="solid")
        fill_card = PatternFill(start_color=COLOR_CARD_BG, end_color=COLOR_CARD_BG, fill_type="solid")
        fill_light_bg = PatternFill(start_color=COLOR_LIGHT_BG, end_color=COLOR_LIGHT_BG, fill_type="solid")

        thin_side = Side(border_style="thin", color=COLOR_BORDER)
        thick_bottom_side = Side(border_style="medium", color=COLOR_PRIMARY)
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom_side)
        border_total = Border(top=Side(border_style="thin", color=COLOR_PRIMARY),
                               bottom=Side(border_style="double", color=COLOR_PRIMARY))

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

        # --- גיליון 1: תקציר מנהלים ---
        ws1 = wb.active
        ws1.title = "תקציר מנהלים וכתב כמויות"
        ws1.views.sheetView[0].rightToLeft = True

        ws1.merge_cells("A1:F1")
        ws1.merge_cells("A2:F2")
        title_cell = ws1["A1"]
        sub_cell = ws1["A2"]
        title_cell.value = "דוח פיקוח הנדסי ובקרה | תשתיות רמזורים"
        title_cell.font = font_main_title
        title_cell.fill = fill_primary
        title_cell.alignment = align_center

        sub_cell.value = f"פרויקט / צומת: {junction_name} | תאריך סיור: {inspection_date}"
        sub_cell.font = font_sub_title
        sub_cell.fill = fill_primary
        sub_cell.alignment = align_center

        ws1.row_dimensions[1].height = 32
        ws1.row_dimensions[2].height = 20

        ws1.cell(row=4, column=1, value="📌 פרטי הסיור והפיקוח").font = font_section_header
        info_data = [
            ("שם הצומת והפרויקט:", junction_name, "מפקח אחראי:", inspector_name),
            ("חברה מבצעת:", contractor_name, "שלב הסדר תנועה:", project_phase),
            ("סוג כבילה מרכזי:", cabling_mode, "סטטוס בטיחות הארקה:",
             "מאושר ותקין" if is_grounding_approved else "❌ לא אושר"),
        ]

        for idx, (lbl1, val1, lbl2, val2) in enumerate(info_data, start=5):
            ws1.cell(row=idx, column=1, value=lbl1).font = font_data_bold
            ws1.cell(row=idx, column=1).fill = fill_light_bg
            ws1.cell(row=idx, column=2, value=val1).font = font_data
            ws1.cell(row=idx, column=4, value=lbl2).font = font_data_bold
            ws1.cell(row=idx, column=4).fill = fill_light_bg
            val2_cell = ws1.cell(row=idx, column=5, value=val2)
            val2_cell.font = font_data
            if "מאושר" in str(val2):
                val2_cell.fill = PatternFill(start_color=COLOR_SUCCESS_BG, fill_type="solid")
                val2_cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_SUCCESS_FG)
            elif "לא אושר" in str(val2):
                val2_cell.fill = PatternFill(start_color=COLOR_DANGER_BG, fill_type="solid")
                val2_cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_DANGER_FG)
            ws1.row_dimensions[idx].height = 22

        ws1.cell(row=8, column=1, value="דגשים והערות המפקח:").font = font_data_bold
        ws1.merge_cells("B8:F8")
        ws1.cell(row=8, column=2, value=general_notes).font = font_data

        ws1.cell(row=10, column=1, value="🔌 אומדן כמויות כבלים מתוכנן").font = font_section_header
        ws1.merge_cells("A11:B11")
        ws1.merge_cells("A12:B12")
        ws1.cell(row=11, column=1, value="כבל פיקוד רמזורים כבד").font = font_metric_lbl
        ws1.cell(row=12, column=1, value=f"{cable_heavy} מ'").font = font_metric_num

        ws1.merge_cells("C11:D11")
        ws1.merge_cells("C12:D12")
        ws1.cell(row=11, column=3, value="כבל פיקוד הולכי רגל").font = font_metric_lbl
        ws1.cell(row=12, column=3, value=f"{cable_light} מ'").font = font_metric_num

        ws1.merge_cells("E11:F11")
        ws1.merge_cells("E12:F12")
        ws1.cell(row=11, column=5, value="כבל הארקה תקני").font = font_metric_lbl
        ws1.cell(row=12, column=5, value=f"{cable_ground} מ'").font = font_metric_num

        for r in [11, 12]:
            for c in range(1, 7):
                cell = ws1.cell(row=r, column=c)
                cell.fill = fill_card
                cell.alignment = align_center
                cell.border = border_all

        ws1.cell(row=14, column=1, value="📋 כתב כמויות ציוד ורכיבי צומת").font = font_section_header
        headers = ["תיאור הרכיב / התשתית", "כמות מצב קיים (לפני)", "כמות מתוכננת (אחרי)",
                   "שינוי (דלתא Δ)", "יחידת מידה", "הערות הנדסיות"]
        ws1.row_dimensions[15].height = 26
        for col_idx, text in enumerate(headers, start=1):
            cell = ws1.cell(row=15, column=col_idx, value=text)
            cell.fill = fill_secondary
            cell.font = font_tbl_header
            cell.alignment = align_center
            cell.border = border_header

        start_row = 16
        for r_idx, row in df_boq.iterrows():
            current_row = start_row + r_idx
            ws1.row_dimensions[current_row].height = 22
            ws1.cell(row=current_row, column=1, value=row["תיאור הרכיב"])
            ws1.cell(row=current_row, column=2, value=row["לפני"])
            ws1.cell(row=current_row, column=3, value=row["אחרי"])
            c4 = ws1.cell(row=current_row, column=4, value=f"=C{current_row}-B{current_row}")
            ws1.cell(row=current_row, column=5, value="יחידות")
            ws1.cell(row=current_row, column=6,
                     value="תקין ומאושר" if row["אחרי"] >= row["לפני"] else "פורק מהשטח")
            row_fill = fill_zebra if r_idx % 2 == 1 else PatternFill(fill_type=None)

            for c_idx in range(1, 7):
                cell = ws1.cell(row=current_row, column=c_idx)
                cell.font = font_data
                cell.border = border_all
                if row_fill.fill_type:
                    cell.fill = row_fill
                if c_idx in [2, 3, 4, 5]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_right
            c4.font = font_data_bold

        total_row = start_row + len(df_boq)
        ws1.row_dimensions[total_row].height = 24
        ws1.cell(row=total_row, column=1, value='סה"כ רכיבים בצומת').font = font_data_bold
        ws1.cell(row=total_row, column=2, value=f"=SUM(B{start_row}:B{total_row - 1})").font = font_data_bold
        ws1.cell(row=total_row, column=3, value=f"=SUM(C{start_row}:C{total_row - 1})").font = font_data_bold
        ws1.cell(row=total_row, column=4, value=f"=SUM(D{start_row}:D{total_row - 1})").font = font_data_bold

        for c_idx in range(1, 7):
            cell = ws1.cell(row=total_row, column=c_idx)
            cell.border = border_total
            cell.fill = fill_light_bg
            if c_idx in [2, 3, 4]:
                cell.alignment = align_center

        # --- גיליון 2: פרוטוקול הארקה ---
        ws2 = wb.create_sheet(title="פרוטוקול הארקה ובטיחות")
        ws2.views.sheetView[0].rightToLeft = True
        ws2.merge_cells("A1:E1")
        t2 = ws2["A1"]
        t2.value = f"פרוטוקול בדיקת הארקה ובטיחות חשמל | {junction_name}"
        t2.font = font_main_title
        t2.fill = fill_primary
        t2.alignment = align_center
        ws2.row_dimensions[1].height = 30

        g_headers = ["סעיף בדיקה", "תיאור הבדיקה הנדרשת", "תוצאה / ערך נמדד", "סטטוס אישור", "הערות"]
        ws2.row_dimensions[3].height = 24
        for c_idx, text in enumerate(g_headers, start=1):
            cell = ws2.cell(row=3, column=c_idx, value=text)
            cell.fill = fill_secondary
            cell.font = font_tbl_header
            cell.alignment = align_center

        g_checks = [
            ("1", "בדיקת חיבור הארקה לכל עמודי הצומת", "תקין", "מאושר" if g_poles else "נפסל",
             "נבדקו כל רגלי העמודים"),
            ("2", "הארקת ארון פיקוד מרכזי", "תקין", "מאושר" if g_cabinet else "נפסל",
             "חיבור ישיר לפס הארקה"),
            ("3", "בדיקת רציפות הארקה (Continuity Test)", "תקין", "מאושר" if g_continuity else "נפסל",
             "בדיקה בלולאה סגורה"),
            ("4", "מדידת התנגדות הארקה לקו (תקן החשמל)", f"{g_value} Ω",
             "מאושר" if g_value <= 5.0 else "חריגה", "ערך תקני מתחת ל-5 אוהם"),
        ]

        for idx, (num, desc, val, stat, comm) in enumerate(g_checks, start=4):
            ws2.row_dimensions[idx].height = 22
            ws2.cell(row=idx, column=1, value=num).alignment = align_center
            ws2.cell(row=idx, column=2, value=desc).alignment = align_right
            ws2.cell(row=idx, column=3, value=val).alignment = align_center
            stat_cell = ws2.cell(row=idx, column=4, value=stat)
            stat_cell.alignment = align_center
            stat_cell.font = font_data_bold
            if stat == "מאושר":
                stat_cell.fill = PatternFill(start_color=COLOR_SUCCESS_BG, fill_type="solid")
                stat_cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_SUCCESS_FG)
            else:
                stat_cell.fill = PatternFill(start_color=COLOR_DANGER_BG, fill_type="solid")
                stat_cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_DANGER_FG)
            ws2.cell(row=idx, column=5, value=comm).alignment = align_right
            for c in range(1, 6):
                ws2.cell(row=idx, column=c).border = border_all

        # --- גיליון 3: סקיצת תוואי ---
        ws3 = wb.create_sheet(title="סקיצת תוואי ומיקומים")
        ws3.views.sheetView[0].rightToLeft = True
        ws3.merge_cells("A1:H1")
        t3 = ws3["A1"]
        t3.value = f"תרשים סקיצה הנדסית ותוואי צומת | {junction_name}"
        t3.font = font_main_title
        t3.fill = fill_primary
        t3.alignment = align_center
        ws3.row_dimensions[1].height = 30

        bg_img_cur = st.session_state.get("active_bg_img", None)
        sketch_data = st.session_state.get("canvas_sketch_image", None)

        if bg_img_cur is not None or sketch_data is not None:
            if bg_img_cur is not None:
                final_combined_img = bg_img_cur.copy().resize((900, 500)).convert("RGB")
            else:
                final_combined_img = Image.new("RGB", (900, 500), (255, 255, 255))

            if sketch_data is not None:
                sketch_img = Image.fromarray(sketch_data.astype('uint8'), 'RGBA')
                final_combined_img.paste(sketch_img, (0, 0), sketch_img)

            img_byte_arr = io.BytesIO()
            final_combined_img.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)

            excel_img = OpenpyxlImage(img_byte_arr)
            excel_img.width = 820
            excel_img.height = 455
            ws3.add_image(excel_img, "B3")

            sk_counts = st.session_state.get("sketch_counts", {})
            ws3.cell(row=30, column=2, value="ספירת אלמנטים שזוהו בסקיצה בפועל").font = font_section_header
            row_cursor = 31
            for k, v in ELEMENT_BANK.items():
                ws3.cell(row=row_cursor, column=2, value=v["label"]).font = font_data
                ws3.cell(row=row_cursor, column=3, value=sk_counts.get(k, 0)).font = font_data_bold
                row_cursor += 1
        else:
            ws3.cell(row=3, column=2, value="לא הופקה סקיצה בטאב הראשון.").font = font_data

        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in [1, 2]:
                        continue
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

        wb.save(output)
        return output.getvalue()

    excel_file = generate_excel_report()
    st.markdown("### 📄 הורדת הקובץ המוגמר")
    st.download_button(
        label="📥 הורד דוח Excel הנדסי מעוצב",
        data=excel_file,
        file_name=f"Junction_Report_{junction_name.replace(' ', '_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
